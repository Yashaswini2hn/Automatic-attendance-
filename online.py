import os
import tkinter as tk
import cv2
from PIL import Image, ImageTk
from datetime import datetime
from mtcnn import MTCNN
import face_recognition
from openpyxl import Workbook
import shutil
from datetime import datetime
import time

class VideoPlayer:
    def __init__(self, master):
        self.master = master
        self.master.title("Online Video Player")
        self.master.geometry("800x500")
        self.master.resizable(False, False)

        # create video frame
        self.video_frame = tk.Frame(self.master, relief=tk.SUNKEN, bd=2)
        self.video_frame.place(x=10, y=10, width=640, height=480)

        # create canvas for displaying video
        self.canvas = tk.Canvas(self.video_frame, bg='black', width=640, height=480)
        self.canvas.pack()

        # create widgets
        self.button_open_camera = tk.Button(self.master, text="Open Camera", command=self.open_camera, borderwidth="10", bg="red", activebackground="green")
        self.button_open_camera.place(x=670, y=200)
        self.button_screenshot = tk.Button(self.master, text="Take Screenshot", command=self.take_screenshot,borderwidth="10",bg="red",activebackground="green")
        self.button_screenshot.place(x=670, y=260)
        self.button_compare_faces = tk.Button(self.master, text="Compare Faces""\n""&Attendance", command=self.compare_faces_and_mark_attendance,bg="red",borderwidth="10",activebackground="green")
        self.button_compare_faces.place(x=670, y=330)
        self.button_compare_faces = tk.Button(self.master, text="Delete screenshot", command=self.delete_old_subfolders,bg="red",borderwidth="10",activebackground="green")
        self.button_compare_faces.place(x=670, y=410)
    
        # initialize video capture
        self.cap = None
        self.detector = MTCNN()
        self.face_encodings = []
        self.dataset_faces = {}
        dataset_path = "dataset"  # Replace this with the path to your dataset folder
        self.load_dataset_faces(dataset_path)

    def open_camera(self):
        if self.cap:
            self.cap.release()
        self.cap = cv2.VideoCapture(1)
        self.play()

    def play(self):
        ret, frame = self.cap.read()
        if ret:
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

            # Detect faces using MTCNN
            faces = self.detector.detect_faces(frame)
            if len(faces) >= 5:
                self.auto_screenshot()
            # Draw rectangles around the detected faces
            for face in faces:
                x, y, w, h = face['box']
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)

                # Find the closest match in the dataset for the detected face
                face_location = (y, x + w, y + h, x)
                face_encoding = face_recognition.face_encodings(frame, [face_location])[0]
                best_match_name = None
                best_match_distance = float("inf")
                for person_name, encodings in self.dataset_faces.items():
                    distances = face_recognition.face_distance(encodings, face_encoding)
                    min_distance = min(distances)
                    if min_distance < best_match_distance:
                        best_match_distance = min_distance
                        best_match_name = person_name

                if best_match_distance < 0.6:
                    cv2.putText(frame, best_match_name, (x, y + h + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)

            frame = cv2.resize(frame, (640, 480))

            image = Image.fromarray(frame)
            photo = ImageTk.PhotoImage(image)
            self.canvas.create_image(0, 0, image=photo, anchor=tk.NW)
            self.canvas.image = photo
            self.master.after(10, self.play)
    def auto_screenshot(self):
        self.detector = MTCNN()
        ret, frame = self.cap.read()
        if ret:
            # convert frame to RGB format
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            faces = self.detector.detect_faces(frame)
            if len(faces) >= 5:
                parent_folder = "screenshot_online"
                today = datetime.now().strftime('%Y-%m-%d')

                if not os.path.exists(parent_folder):
                    os.makedirs(parent_folder)

                subfolder_path = os.path.join(parent_folder, today)

                if not os.path.exists(subfolder_path):
                    os.makedirs(subfolder_path)

                # convert frame back to BGR format
                frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)

                # save frame as PNG file
                filename = datetime.now().strftime("%Y-%m-%d_%H-%M-%S_autoscreenshot.png")
                # Save the screenshot to the "screenshots" folder
                filepath = os.path.join(subfolder_path, filename)
                cv2.imwrite(filepath, frame)

    def load_dataset_faces(self, dataset_path):
        for person_name in os.listdir(dataset_path):
            person_dir = os.path.join(dataset_path, person_name)
            if os.path.isdir(person_dir):
                for filename in os.listdir(person_dir):
                    if filename.endswith('.jpg') or filename.endswith('.png'):
                        image_path = os.path.join(person_dir, filename)
                        face_image = face_recognition.load_image_file(image_path)
                        face_locations = face_recognition.face_locations(face_image)

                        if len(face_locations) > 0:
                            face_encoding = face_recognition.face_encodings(face_image, face_locations)[0]
                            if person_name not in self.dataset_faces:
                                self.dataset_faces[person_name] = []
                            self.dataset_faces[person_name].append(face_encoding)
    
    def delete_old_subfolders(self, parent_folder="screenshot_online", days_old=7):
            current_time = time.time()
            for folder in os.listdir(parent_folder):
                folder_path = os.path.join(parent_folder, folder)
                if os.path.isdir(folder_path):
                    folder_creation_time = os.path.getctime(folder_path)
                    folder_age = current_time - folder_creation_time

                    if folder_age > days_old * 24 * 60 * 60:  # Convert days to seconds
                        shutil.rmtree(folder_path)
                        print(f"Deleted folder: {folder_path}")


    def take_screenshot(self):
        # read current frame from video capture
        ret, frame = self.cap.read()
        if ret:
            # convert frame to RGB format
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

            parent_folder = "screenshot_online"
            today = datetime.now().strftime('%Y-%m-%d')

            if not os.path.exists(parent_folder):
                os.makedirs(parent_folder)

            subfolder_path = os.path.join(parent_folder, today)

            if not os.path.exists(subfolder_path):
                os.makedirs(subfolder_path)

            # convert frame back to BGR format
            frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)

            # save frame as PNG file
            filename = datetime.now().strftime("%Y-%m-%d_%H-%M-%S_screenshot.png")
            # Save the screenshot to the "screenshots" folder
            filepath = os.path.join(subfolder_path, filename)
            cv2.imwrite(filepath, frame)

    def compare_faces_and_mark_attendance(self):
        dataset_path = "dataset"
        parent_folder = "screenshot_online"
        today = datetime.now().strftime('%Y-%m-%d')
        subfolder_path = os.path.join(parent_folder, today)
        
        if dataset_path:
            self.load_dataset_faces(dataset_path)

        attendance = {}

        for file in os.listdir(subfolder_path):
            file_path = os.path.join(subfolder_path, file)
            if os.path.isfile(file_path):
                frame = cv2.imread(file_path)

                # convert frame to RGB format
                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

                # Detect faces using MTCNN
                faces = self.detector.detect_faces(frame)

                # Find the closest match in the dataset for the detected face
                for face in faces:
                    x, y, w, h = face['box']
                    face_location = (y, x + w, y + h, x)
                    face_encoding = face_recognition.face_encodings(frame, [face_location])[0]
                    best_match_name = None
                    best_match_distance = float("inf")
                    for person_name, encodings in self.dataset_faces.items():
                        distances = face_recognition.face_distance(encodings, face_encoding)
                        min_distance = min(distances)
                        if min_distance < best_match_distance:
                            best_match_distance = min_distance
                            best_match_name = person_name

                    if best_match_distance < 0.6:
                        attendance[best_match_name] = datetime.now()

        teacher_name, section, subject_code,sem = self.show_input_dialog()
        self.create_excel_attendance(teacher_name, section, subject_code,sem, attendance)

    def show_input_dialog(self):
        dialog = tk.Toplevel(self.master)
        dialog.title("Enter Details")
        dialog.geometry("300x200")

        teacher_name_label = tk.Label(dialog, text="Teacher Name:")
        teacher_name_label.grid(row=0, column=0, padx=10, pady=10)
        teacher_name_entry = tk.Entry(dialog)
        teacher_name_entry.grid(row=0, column=1)

        section_label = tk.Label(dialog, text="Section:")
        section_label.grid(row=1, column=0, padx=10, pady=10)
        section_entry = tk.Entry(dialog)
        section_entry.grid(row=1, column=1)

        subject_code_label = tk.Label(dialog, text="Subject_code:")
        subject_code_label.grid(row=2, column=0, padx=10, pady=10)
        subject_code_entry = tk.Entry(dialog)
        subject_code_entry.grid(row=2, column=1)

        sem_label = tk.Label(dialog, text="Sem:")
        sem_label.grid(row=3, column=0, padx=10, pady=10)
        sem_entry = tk.Entry(dialog)
        sem_entry.grid(row=3, column=1)

        # Create variables to store the user input
        result = {'teacher_name': '', 'section': '', 'subject_code': '','sem':'',}

        def submit():
            # Save the user input before destroying the dialog
            result['teacher_name'] = teacher_name_entry.get()
            result['section'] = section_entry.get()
            result['subject_code'] = subject_code_entry.get()
            result['sem']=sem_entry.get()
            dialog.destroy()

        submit_button = tk.Button(dialog, text="Submit", command=submit,borderwidth="10",background="red",activebackground="green")
        submit_button.grid(row=4, column=1, pady=10)

        # Wait for the dialog to close
        self.master.wait_window(dialog)

        # Return the user input
        return result['teacher_name'], result['section'], result['subject_code'],result['sem']


    def create_excel_attendance(self, teacher_name, section, subject_code,sem, attendance_data):
        workbook = Workbook()
        sheet = workbook.active
        dataset_path = "dataset"
        sheet.title = "Attendance"
        all_students = []
        
        for student_name in os.listdir(dataset_path):
            student_path = os.path.join(dataset_path, student_name)
            if os.path.isdir(student_path):
                all_students.append(student_name)
        sheet.cell(row=1, column=1).value = "Teacher Name"
        sheet.cell(row=2, column=1).value = teacher_name
        sheet.cell(row=1, column=2).value = "Section"
        sheet.cell(row=2, column=2).value = section
        sheet.cell(row=1, column=3).value = "Subject_code"
        sheet.cell(row=2, column=3).value = subject_code
        sheet.cell(row=1, column=4).value = "Sem"
        sheet.cell(row=2, column=4).value = sem
        sheet.cell(row=1, column=5).value = "Attendance"

        sheet.cell(row=1, column=6).value = "Name"
        sheet.cell(row=1, column=7).value = "Date & Time"

        for idx, (name, date_time) in enumerate(attendance_data.items(), start=2):
            sheet.cell(row=idx, column=5).value = name
            sheet.cell(row=idx, column=6).value = "Present"
            sheet.cell(row=idx, column=7).value = date_time.strftime('%Y-%m-%d %H:%M:%S')

        present_students = list(attendance_data.keys())
        absent_students = [student for student in all_students if student not in present_students]

        for idx, name in enumerate(absent_students, start=len(attendance_data) + 2):
            sheet.cell(row=idx, column=5).value = name
            sheet.cell(row=idx, column=6).value = "Absent"
            sheet.cell(row=idx, column=7).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')


        
            # Create a folder based on the section name if it doesn't exist
        section_folder = f"Section_{section}"
        if not os.path.exists(section_folder):
            os.makedirs(section_folder)

        # Include the section and subject in the filename
        attendance_filename = datetime.now().strftime(f"{section}_{subject_code}online_%Y-%m-%d_%H-%M-%S.xlsx")
        
        # Save the attendance file inside the section folder
        workbook.save(os.path.join(section_folder, attendance_filename))

if __name__ == '__main__':
    root = tk.Tk()
    app = VideoPlayer(root)
    root.mainloop()
